from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"

INPUT_COLUMNS = [
    "SKU",
    "产品名称",
    "所属类目",
    "采购成本",
    "境内段运费",
    "包装费",
    "其他费用",
    "重量g",
    "长cm",
    "宽cm",
    "高cm",
    "货值区间",
    "物流时效",
    "当前售价人民币",
    "目标利润率",
    "汇率",
    "广告费比例",
    "提现手续费费率",
    "退货率",
    "税费比例",
    "CPC卢布",
    "CPC人民币",
    "预计点击转化率",
    "测试点击数",
    "PPC每周预算卢布",
    "CPO模式",
    "CPO费率",
    "是否组合推广",
]

LOGISTICS_PRICE_ROWS = [
    ["Extra Small", "5-10天", "CEL Express Extra Small", "0.0468 元/克 + 3.12 元/票", 0.0468, 3.12, "超级轻小件（陆空特快）"],
    ["Extra Small", "10-15天", "CEL Standard Extra Small", "0.0364 元/克 + 3.12 元/票", 0.0364, 3.12, "超级轻小件（陆空标准）"],
    ["Extra Small", "15-25天", "CEL Economy Extra Small", "0.026 元/克 + 3.12 元/票", 0.026, 3.12, "超级轻小件（陆运经济）"],
    ["Budget", "5-10天", "CEL Express Budget", "0.03432 元/克 + 23.92 元/票", 0.03432, 23.92, "低客单价标准件（陆空特快）"],
    ["Budget", "10-15天", "CEL Standard Budget", "0.026 元/克 + 23.92 元/票", 0.026, 23.92, "低客单价标准件（陆空标准）"],
    ["Budget", "15-25天", "CEL Economy Budget", "0.01768 元/克 + 23.92 元/票", 0.01768, 23.92, "低客单价标准件（陆运经济）"],
    ["Small", "5-10天", "CEL Express Small", "0.0468 元/克 + 16.64 元/票", 0.0468, 16.64, "小件（陆空特快）"],
    ["Small", "10-15天", "CEL Standard Small", "0.0364 元/克 + 16.64 元/票", 0.0364, 16.64, "小件（陆空标准）"],
    ["Small", "15-25天", "CEL Economy Small", "0.026 元/克 + 16.64 元/票", 0.026, 16.64, "小件（陆运经济）"],
    ["Big", "10-15天", "CEL Standard Big", "0.026 元/克 + 37.44 元/票", 0.026, 37.44, "大件（陆空标准）"],
    ["Big", "15-25天", "CEL Economy Big", "0.01768 元/克 + 37.44 元/票", 0.01768, 37.44, "大件（陆运经济）"],
    ["Premium Small", "5-10天", "CEL Express Premium Small", "0.0468 元/克 + 22.88 元/票", 0.0468, 22.88, "高客单价小件（陆空特快）"],
    ["Premium Small", "10-15天", "CEL Standard Premium Small", "0.0364 元/克 + 22.88 元/票", 0.0364, 22.88, "高客单价小件（陆空标准）"],
    ["Premium Small", "15-25天", "CEL Economy Premium Small", "0.026 元/克 + 22.88 元/票", 0.026, 22.88, "高客单价小件（陆运经济）"],
    ["Premium Big", "10-15天", "CEL Standard Premium Big", "0.02912 元/克 + 64.48 元/票", 0.02912, 64.48, "高客单价大件（陆空标准）"],
    ["Premium Big", "15-25天", "CEL Economy Premium Big", "0.02392 元/克 + 64.48 元/票", 0.02392, 64.48, "高客单价大件（陆运经济）"],
]

EXAMPLE_ROWS = [
    {
        "SKU": "JZ-001",
        "产品名称": "墙面修补工具套装",
        "所属类目": "建筑和装修",
        "采购成本": 28,
        "境内段运费": 3,
        "包装费": 1.5,
        "其他费用": 0,
        "重量g": 420,
        "长cm": 24,
        "宽cm": 16,
        "高cm": 8,
        "货值区间": "1-1500₽",
        "物流时效": "10-15天",
        "当前售价人民币": 95,
        "目标利润率": 0.2,
        "汇率": 12.5,
        "广告费比例": 0,
        "提现手续费费率": 0.01,
        "退货率": 0,
        "税费比例": 0,
        "CPC卢布": 5,
        "CPC人民币": 0,
        "预计点击转化率": 0.02,
        "测试点击数": 30,
        "PPC每周预算卢布": 2000,
        "CPO模式": "全部商品CPO",
        "CPO费率": "",
        "是否组合推广": "否",
    },
    {
        "SKU": "JZ-002",
        "产品名称": "瓷砖找平器",
        "所属类目": "建筑和装修",
        "采购成本": 55,
        "境内段运费": 3,
        "包装费": 2,
        "其他费用": 0,
        "重量g": 1200,
        "长cm": 32,
        "宽cm": 22,
        "高cm": 12,
        "货值区间": "1501-7000₽",
        "物流时效": "10-15天",
        "当前售价人民币": 0,
        "目标利润率": 0.25,
        "汇率": 12.5,
        "广告费比例": 0.05,
        "提现手续费费率": 0.01,
        "退货率": 0.02,
        "税费比例": 0,
        "CPC卢布": 6,
        "CPC人民币": 0,
        "预计点击转化率": 0.03,
        "测试点击数": 50,
        "PPC每周预算卢布": 2000,
        "CPO模式": "PPC+CPO组合",
        "CPO费率": 0.10,
        "是否组合推广": "是",
    },
    {
        "SKU": "JZ-003",
        "产品名称": "重型门窗五金套件",
        "所属类目": "建筑和装修",
        "采购成本": 260,
        "境内段运费": 3,
        "包装费": 8,
        "其他费用": 5,
        "重量g": 6500,
        "长cm": 70,
        "宽cm": 30,
        "高cm": 18,
        "货值区间": "7001-250000₽",
        "物流时效": "15-25天",
        "当前售价人民币": 520,
        "目标利润率": 0.2,
        "汇率": 12.5,
        "广告费比例": 0.03,
        "提现手续费费率": 0.01,
        "退货率": 0.01,
        "税费比例": 0,
        "CPC卢布": 8,
        "CPC人民币": 0,
        "预计点击转化率": 0.02,
        "测试点击数": 50,
        "PPC每周预算卢布": 3000,
        "CPO模式": "单个产品CPO",
        "CPO费率": "",
        "是否组合推广": "否",
    },
]


def build_logistics_prices() -> pd.DataFrame:
    return pd.DataFrame(
        LOGISTICS_PRICE_ROWS,
        columns=["物流标准", "物流时效", "渠道名称", "资费标准", "每克单价", "每票固定费", "备注"],
    )


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="246BFD")
        for idx, column_cells in enumerate(ws.columns, start=1):
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 4, 12), 24)
    wb.save(path)


def write_excel(path: Path, df: pd.DataFrame, sheet_name: str) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    style_workbook(path)


def main() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    build_logistics_prices().to_csv(DATA_DIR / "logistics_prices.csv", index=False, encoding="utf-8-sig")
    write_excel(DATA_DIR / "Ozon批量导入模板.xlsx", pd.DataFrame(columns=INPUT_COLUMNS), "批量导入模板")
    write_excel(DATA_DIR / "Ozon示例数据.xlsx", pd.DataFrame(EXAMPLE_ROWS), "示例数据")


if __name__ == "__main__":
    main()
